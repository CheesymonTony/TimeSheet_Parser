import os
import openpyxl
from openpyxl.styles import PatternFill, Border, Side

class Table:
    def __init__(self, worksheet, headers):
        self.worksheet = worksheet
        self.headers = headers
        self.data = []

    def mergeHeader():
        self.worksheet.merge_cells(self.headers)

newTable = Table


# Create a new Excel workbook
wb = openpyxl.Workbook()

# Select the first sheet
sheet = wb.active
newTable = Table(sheet, ['Header1', 'Header2', 'Header3'])


# Set the column widths for columns B-F
column_widths = [15, 15, 15, 15, 15]
for i, width in enumerate(column_widths):
    sheet.column_dimensions[openpyxl.utils.get_column_letter(i+2)].width = width

# Write your name to cell A1
sheet['A1'] = 'Carter Boyce'

# Write the billing period month to cell A2
billing_period_month = 'June 2022' # Replace with the correct billing period month
sheet['A2'] = billing_period_month

# Insert a blank row after the billing period cell
sheet.insert_rows(3)

# Write the headers to the first row
header_row = sheet['B4':'F4'][0]
header_row[0].value = 'Date'
header_row[1].value = 'In Time'
header_row[2].value = 'Lunch/Break'
header_row[3].value = 'Out Time'
header_row[4].value = 'Additional Hours'
header_fill = PatternFill(start_color='E9B38A', end_color='E9B38A', fill_type='solid')
for cell in header_row:
    cell.fill = header_fill

# Write the data to the appropriate cells
data_fill = PatternFill(start_color='F7E4D7', end_color='F7E4D7', fill_type='solid')
for i in range(5, 32):
    # Write the day of the month to column B
    sheet.cell(row=i, column=2, value=i-4).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Write the rest of the data to columns C-F
    sheet.cell(row=i, column=3, value='').border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    sheet.cell(row=i, column=4, value='').border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    sheet.cell(row=i, column=5, value='').border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    sheet.cell(row=i, column=6, value='').border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for j in range(2, 7):
        sheet.cell(row=i, column=j).fill = data_fill

# Set border around table
thin_border = openpyxl.styles.Side(border_style='thin', color='000000')
thick_border = openpyxl.styles.Side(border_style='medium', color='000000')
border = openpyxl.styles.Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)

# Set border around table headers\
print('header_row',len(header_row))
for idx, cell in enumerate(header_row):
    cell.fill = openpyxl.styles.PatternFill(start_color='E9B38A', end_color='E9B38A', fill_type='solid')
    if idx == 0:
        cell.border = openpyxl.styles.Border(top=thick_border, left=thick_border, right=thin_border, bottom=thick_border)
    elif idx == len(header_row) - 1:
        cell.border = openpyxl.styles.Border(top=thick_border, left=thin_border, right=thick_border, bottom=thick_border)
    else:
        cell.border = openpyxl.styles.Border(top=thick_border, left=thin_border, right=thin_border, bottom=thick_border)

            

# Set border around table cells
# for row in sheet.iter_rows(min_row=5, max_row=31, min_col=2, max_col=6):
#     for cell in row:
#         cell.fill = openpyxl.styles.PatternFill(start_color='F7E4D7', end_color='F7E4D7', fill_type='solid')
#         cell.border = openpyxl.styles.Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)
        
# Set border around entire table
min_col = sheet.min_column
max_col = sheet.max_column
min_row = sheet.min_row
max_row = sheet.max_row

# for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
#     for cell in row:
#         if cell.row == min_row or cell.row == max_row or cell.column == min_col or cell.column == max_col:
#             cell.border = openpyxl.styles.Border(top=thick_border, left=thick_border, right=thick_border, bottom=thick_border)


# Save the Excel workbook in the current directory
file_path = os.path.join(os.getcwd(), 'example.xlsx')
wb.save(file_path)
